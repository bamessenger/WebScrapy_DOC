from openpyxl import Workbook, load_workbook


class WriteFile:
    def __init__(self):
        self.wb = Workbook()
        self.existingDict = {}

    def currExcel2Dict(self, currMaster):
        # load workbook
        self.wb = load_workbook(currMaster)
        # select workbook
        sheet = self.wb.active
        # get max row count
        max_row = sheet.max_row
        # iterate through all rows in first column, skipping first row
        for r in range(2, max_row + 1):
            key = sheet.cell(row=r, column=1)
            entityName = sheet.cell(row=r, column=2)
            incDate = sheet.cell(row=r, column=3)
            self.existingDict[key.value] = entityName.value, incDate.value
        self.wb.close()

    def dictMerge(self, dict2):
        self.existingDict.update(dict2)

    def newMast2Excel(self, currMaster):
        # load workbook
        self.wb = load_workbook(currMaster)
        # select workbook
        sheet = self.wb.active
        # re-establish headers
        sheet.cell(row=1, column=1).value = "File Number"
        sheet.cell(row=1, column=2).value = "Entity Name"
        sheet.cell(row=1, column=3).value = "Inc Date"
        # iterate through all rows in first column, skipping first row
        r = 1
        for key in self.existingDict.keys():
            c = 2
            r += 1
            sheet.cell(row=r, column=1).value = key
            for val in self.existingDict[key]:
                sheet.cell(row=r, column=c).value = val
                c += 1
        self.wb.save(currMaster)
        self.wb.close()

    def setNewMastDict(self, newDict, mFile):
        # grab current excel data and create a temporary dictionary
        self.currExcel2Dict(mFile)
        # receive scraped data dictionary and pass it to merging method
        self.dictMerge(newDict)
        # once both dictionaries are merged, re-write the updated dictionary
        # back to Excel file
        self.newMast2Excel(mFile)
