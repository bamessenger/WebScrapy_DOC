from openpyxl import Workbook, load_workbook


class ParseFile:
    def __init__(self, fPath):
        self.filepath = fPath
        self.wb = Workbook()
        self.search_list = []
    def fileParser(self):
        # load workbook
        self.wb = load_workbook(self.filepath)
        # select workbook
        sheet = self.wb.active
        # get max row count
        max_row = sheet.max_row
        # iterate through all rows in first column, skipping first row
        for r in range(2, max_row + 1):
            cell = sheet.cell(row=r, column=1)
            self.search_list.append(cell.value)

    def getExcelFile(self):
        self.fileParser()
        return self.search_list